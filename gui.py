import sys
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtCore import *
from PyQt5.QtWidgets import QApplication, QDialog
from PyQt5.uic import loadUi
import cv2
import numpy as np
from scipy.spatial import distance as dist
import imutils
import pyautogui
#import win32api, win32con
#from win32api import GetSystemMetrics

from openpyxl import load_workbook

class ColorDetector(QDialog):
    frame_counter = 0
    def __init__(self):
        super(ColorDetector, self).__init__()
        loadUi('tracking.ui', self)

        self.file_name = 'data.xlsx'
        self.wb = load_workbook(filename = self.file_name)
        self.ws = self.wb.active
        
        self.screenWidth = GetSystemMetrics(0)
        self.screenHeight = GetSystemMetrics(1)
        self.screenOrigin = 1366
        self.screenExtends = self.screenWidth - self.screenOrigin

        self.image = None
        self.mode = "corner"
        self.mouse_track = False
        self.mouse_track_status.setText("idle")
        self.count = 0 

        self.start_button.clicked.connect(self.start_webcam)
        self.stop_button.clicked.connect(self.stop_webcam)
        self.save_corner.clicked.connect(self.set_corner)
        self.save_pointer.clicked.connect(self.set_pointer)
        self.load_corner.clicked.connect(self.load_data_corner)
        self.load_pointer.clicked.connect(self.load_data_pointer)
        self.track_button.clicked.connect(self.start_mouse_track)
        self.stop_track_button.clicked.connect(self.stop_mouse_track)
        self.c_up.clicked.connect(self.push_up)
        self.c_down.clicked.connect(self.push_down)
        self.c_left.clicked.connect(self.push_left)
        self.c_right.clicked.connect(self.push_right)
        self.corner_adj_save.clicked.connect(self.push_save)
        self.pushButton_reset.clicked.connect(self.push_reset)

        self.ca1 = int(self.ws['A1'].value)
        self.ca2 = int(self.ws['A2'].value)
        self.ca3 = int(self.ws['A3'].value)

        self.cb1 = int(self.ws['B1'].value)
        self.cb2 = int(self.ws['B2'].value)
        self.cb3 = int(self.ws['B3'].value)

        self.cc1 = int(self.ws['C1'].value)
        self.cc2 = int(self.ws['C2'].value)
        self.cc3 = int(self.ws['C3'].value)

        self.cd1 = int(self.ws['D1'].value)
        self.cd2 = int(self.ws['D2'].value)
        self.cd3 = int(self.ws['D3'].value)

        self.ltx_adj = int(self.ws['E1'].value)
        self.lty_adj = int(self.ws['F1'].value)

        self.rtx_adj = int(self.ws['E2'].value)
        self.rty_adj = int(self.ws['F2'].value)

        self.brx_adj = int(self.ws['E3'].value)
        self.bry_adj = int(self.ws['F3'].value)

        self.blx_adj = int(self.ws['E4'].value)
        self.bly_adj = int(self.ws['F4'].value)


        self.corner_lower = np.array([self.ca1,self.ca2,self.ca3],np.uint8)
        self.corner_upper = np.array([self.cb1,self.cb2,self.cb3],np.uint8)

        self.pointer_lower = np.array([self.cc1,self.cc2,self.cc3],np.uint8)
        self.pointer_upper = np.array([self.cd1,self.cd2,self.cd3],np.uint8)

        self.load_data_corner()

        self.h_min_val.setText(str(self.h_min.value()))
        self.s_min_val.setText(str(self.s_min.value()))
        self.v_min_val.setText(str(self.v_min.value()))
        self.h_max_val.setText(str(self.h_max.value()))
        self.s_max_val.setText(str(self.s_max.value()))
        self.v_max_val.setText(str(self.v_max.value()))

        self.h_min_corner.setText(str(self.corner_lower[0]))
        self.s_min_corner.setText(str(self.corner_lower[1]))
        self.v_min_corner.setText(str(self.corner_lower[2]))

        self.h_max_corner.setText(str(self.corner_upper[0]))
        self.s_max_corner.setText(str(self.corner_upper[1]))
        self.v_max_corner.setText(str(self.corner_upper[2]))

        self.h_min_pointer.setText(str(self.pointer_lower[0]))
        self.s_min_pointer.setText(str(self.pointer_lower[1]))
        self.v_min_pointer.setText(str(self.pointer_lower[2]))

        self.h_max_pointer.setText(str(self.pointer_upper[0]))
        self.s_max_pointer.setText(str(self.pointer_upper[1]))
        self.v_max_pointer.setText(str(self.pointer_upper[2]))

        self.a1 = 0,0
        self.a2 = 0,0
        self.a3 = 0,0
        self.a4 = 0,0

        self.h = None

    def click(x,y):
        win32api.SetCursorPos((x,y))
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,x,y,0,0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,x,y,0,0)    
        

    def push_reset(self):

        self.ltx_adj = 0
        self.rtx_adj = 0
        self.brx_adj = 0
        self.blx_adj = 0

        self.lty_adj = 0
        self.rty_adj = 0
        self.bry_adj = 0
        self.bly_adj = 0

        self.ws['E1'] = self.ltx_adj
        self.ws['E2'] = self.rtx_adj
        self.ws['E3'] = self.brx_adj
        self.ws['E4'] = self.blx_adj

        self.ws['F1'] = self.lty_adj
        self.ws['F2'] = self.rty_adj
        self.ws['F3'] = self.bry_adj
        self.ws['F4'] = self.bly_adj

        self.wb.save(self.file_name)

    def push_save(self):
        self.ws['E1'] = self.ltx_adj
        self.ws['E2'] = self.rtx_adj
        self.ws['E3'] = self.brx_adj
        self.ws['E4'] = self.blx_adj

        self.ws['F1'] = self.lty_adj
        self.ws['F2'] = self.rty_adj
        self.ws['F3'] = self.bry_adj
        self.ws['F4'] = self.bly_adj

        self.wb.save(self.file_name)

    def push_up(self):
        if self.corner_select.currentText() == "C1":
            self.lty_adj -= int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C2":
            self.rty_adj -= int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C3":
            self.bry_adj -= int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C4":
            self.bly_adj -= int(self.corner_adj.currentText())

        self.corner_adj_func()

    def push_down(self):
        if self.corner_select.currentText() == "C1":
            self.lty_adj += int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C2":
            self.rty_adj += int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C3":
            self.bry_adj += int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C4":
            self.bly_adj += int(self.corner_adj.currentText())
        
        self.corner_adj_func()

    def push_left(self):
        if self.corner_select.currentText() == "C1":
            self.ltx_adj -= int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C2":
            self.rtx_adj -= int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C3":
            self.brx_adj -= int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C4":
            self.blx_adj -= int(self.corner_adj.currentText())
        
        self.corner_adj_func()

    def push_right(self):
        if self.corner_select.currentText() == "C1":
            self.ltx_adj += int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C2":
            self.rtx_adj += int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C3":
            self.brx_adj += int(self.corner_adj.currentText())
        if self.corner_select.currentText() == "C4":
            self.blx_adj += int(self.corner_adj.currentText())
        
        self.corner_adj_func()

    def start_mouse_track(self):
        self.mouse_track = True
        self.mouse_track_status.setText("Tracking") 

    def stop_mouse_track(self):
        self.mouse_track = False
        self.mouse_track_status.setText("idle") 

    def load_data_corner(self):
        self.mode = "corner"
        self.mode_view.setText(self.mode)
        self.h_min.setValue(self.corner_lower[0]) 
        self.s_min.setValue(self.corner_lower[1])
        self.v_min.setValue(self.corner_lower[2])

        self.h_max.setValue(self.corner_upper[0])
        self.s_max.setValue(self.corner_upper[1])
        self.v_max.setValue(self.corner_upper[2])

    def load_data_pointer(self):
        self.mode = "pointer"
        self.mode_view.setText(self.mode)
        self.h_min.setValue(self.pointer_lower[0])
        self.s_min.setValue(self.pointer_lower[1])
        self.v_min.setValue(self.pointer_lower[2])

        self.h_max.setValue(self.pointer_upper[0])
        self.s_max.setValue(self.pointer_upper[1])
        self.v_max.setValue(self.pointer_upper[2])

    def corner_adj_func(self):

        erode = cv2.erode(self.color_mask, None, iterations=2)
        dilate = cv2.dilate(erode,None, iterations=10)
        
        kernelOpen = np.ones((5,5))
        kernelClose = np.ones((20,20))
        
        maskOpen = cv2.morphologyEx(dilate, cv2.MORPH_OPEN, kernelOpen)
        maskClose = cv2.morphologyEx(maskOpen, cv2.MORPH_CLOSE, kernelClose)
     
        corners = cv2.goodFeaturesToTrack(maskClose,4,0.01,150)

        cx = [0,0,0,0]
        cy = [0,0,0,0]

        i_count = 0

        if np.any(corners == None):
            pass

        else:
            
            corners = np.int0(corners)

            for i in corners:
                x,y = i.ravel()
                #cv2.circle(img,(x,y),5,(0, 0, 255),-1)
                cx[i_count]= x
                cy[i_count]= y
                i_count += 1

        #img = cv2.drawContours(img, contours, -1, (0, 255, 0), 2)

        c1 = cx[0],cy[0]
        c2 = cx[1],cy[1]
        c3 = cx[2],cy[2]
        c4 = cx[3],cy[3]
       
        cs = [0,0,0,0]

        for i in range(4):
            cs[i] = cx[i] + cy[i]

        csmin = cs[0]
        csmax = cs[0]

        ltx = cx[0]
        lty = cy[0]

        brx = cx[0]
        bry = cy[0]

        for i in range(4):
            if csmin > cs[i]:
                csmin = cs[i]
                ltx = cx[i]
                lty = cy[i]
        
        for i in range(4):
            if csmax < cs[i]:
                csmax = cs[i]
                brx = cx[i]
                bry = cy[i]
        
        cd = [0,0,0,0]

        for i in range(4):
            cd[i] = cx[i] - cy[i]

        cdmin = cd[0]
        cdmax = cd[0]

        rtx = cx[0]
        rty = cy[0]

        blx = cx[0]
        bly = cy[0]


        for i in range(4):
            if cdmin > cd[i]:
                cdmin = cd[i]
                blx = cx[i]
                bly = cy[i]
        
        for i in range(4):
            if cdmax < cd[i]:
                cdmax = cd[i]
                rtx = cx[i]
                rty = cy[i]

        ltx += self.ltx_adj
        lty += self.lty_adj

        rtx += self.rtx_adj
        rty += self.rty_adj

        brx += self.brx_adj
        bry += self.bry_adj

        blx += self.blx_adj
        bly += self.bly_adj

        self.a1 = ltx,lty
        self.a2 = rtx,rty
        self.a3 = brx,bry
        self.a4 = blx,bly

        pts_src = np.array([[ltx, lty], [rtx, rty], [brx, bry],[blx, bly]])
        pts_dst = np.array([[0, 0],[self.screenExtends, 0],[self.screenExtends, self.screenHeight],[0, self.screenHeight]])
        self.h, status = cv2.findHomography(pts_src, pts_dst)

    def set_corner(self):
        self.corner_lower = np.array([self.h_min.value(), self.s_min.value(), self.v_min.value()],np.uint8)
        self.corner_upper = np.array([self.h_max.value(), self.s_max.value(), self.v_max.value()],np.uint8)

        self.h_min_corner.setText(str(self.corner_lower[0]))
        self.s_min_corner.setText(str(self.corner_lower[1]))
        self.v_min_corner.setText(str(self.corner_lower[2]))

        self.h_max_corner.setText(str(self.corner_upper[0]))
        self.s_max_corner.setText(str(self.corner_upper[1]))
        self.v_max_corner.setText(str(self.corner_upper[2]))

        self.ws['A1'] = self.corner_lower[0]
        self.ws['A2'] = self.corner_lower[1]
        self.ws['A3'] = self.corner_lower[2]

        self.ws['B1'] = self.corner_upper[0]
        self.ws['B2'] = self.corner_upper[1]
        self.ws['B3'] = self.corner_upper[2]

        self.wb.save(self.file_name)

        erode = cv2.erode(self.color_mask, None, iterations=2)
        dilate = cv2.dilate(erode,None, iterations=10)
        
        kernelOpen = np.ones((5,5))
        kernelClose = np.ones((20,20))
        
        maskOpen = cv2.morphologyEx(dilate, cv2.MORPH_OPEN, kernelOpen)
        maskClose = cv2.morphologyEx(maskOpen, cv2.MORPH_CLOSE, kernelClose)
     
        corners = cv2.goodFeaturesToTrack(maskClose,4,0.01,150)

        cx = [0,0,0,0]
        cy = [0,0,0,0]

        i_count = 0

        if np.any(corners == None):
            pass

        else:
            
            corners = np.int0(corners)

            for i in corners:
                x,y = i.ravel()
                #cv2.circle(img,(x,y),5,(0, 0, 255),-1)
                cx[i_count]= x
                cy[i_count]= y
                i_count += 1

        #img = cv2.drawContours(img, contours, -1, (0, 255, 0), 2)

        c1 = cx[0],cy[0]
        c2 = cx[1],cy[1]
        c3 = cx[2],cy[2]
        c4 = cx[3],cy[3]
       
        cs = [0,0,0,0]

        for i in range(4):
            cs[i] = cx[i] + cy[i]

        csmin = cs[0]
        csmax = cs[0]

        ltx = cx[0]
        lty = cy[0]

        brx = cx[0]
        bry = cy[0]

        for i in range(4):
            if csmin > cs[i]:
                csmin = cs[i]
                ltx = cx[i]
                lty = cy[i]
        
        for i in range(4):
            if csmax < cs[i]:
                csmax = cs[i]
                brx = cx[i]
                bry = cy[i]
        
        cd = [0,0,0,0]

        for i in range(4):
            cd[i] = cx[i] - cy[i]

        cdmin = cd[0]
        cdmax = cd[0]

        rtx = cx[0]
        rty = cy[0]

        blx = cx[0]
        bly = cy[0]


        for i in range(4):
            if cdmin > cd[i]:
                cdmin = cd[i]
                blx = cx[i]
                bly = cy[i]
        
        for i in range(4):
            if cdmax < cd[i]:
                cdmax = cd[i]
                rtx = cx[i]
                rty = cy[i]

        ltx += self.ltx_adj
        lty += self.lty_adj

        rtx += self.rtx_adj
        rty += self.rty_adj

        brx += self.brx_adj
        bry += self.bry_adj

        blx += self.blx_adj
        bly += self.bly_adj

        self.a1 = ltx,lty
        self.a2 = rtx,rty
        self.a3 = brx,bry
        self.a4 = blx,bly

        pts_src = np.array([[ltx, lty], [rtx, rty], [brx, bry],[blx, bly]])
        #pts_dst = np.array([[0, 0],[1023, 0],[1023, 768],[0, 768]])
        pts_dst = np.array([[0, 0],[self.screenExtends, 0],[self.screenExtends, self.screenHeight],[0, self.screenHeight]])
        self.h, status = cv2.findHomography(pts_src, pts_dst)

    def set_pointer(self):
        self.pointer_lower = np.array([self.h_min.value(), self.s_min.value(), self.v_min.value()],np.uint8)
        self.pointer_upper = np.array([self.h_max.value(), self.s_max.value(), self.v_max.value()],np.uint8)

        self.h_min_pointer.setText(str(self.pointer_lower[0]))
        self.s_min_pointer.setText(str(self.pointer_lower[1]))
        self.v_min_pointer.setText(str(self.pointer_lower[2]))

        self.h_max_pointer.setText(str(self.pointer_upper[0]))
        self.s_max_pointer.setText(str(self.pointer_upper[1]))
        self.v_max_pointer.setText(str(self.pointer_upper[2]))

        self.ws['C1'] = self.pointer_lower[0]
        self.ws['C2'] = self.pointer_lower[1]
        self.ws['C3'] = self.pointer_lower[2]

        self.ws['D1'] = self.pointer_upper[0]
        self.ws['D2'] = self.pointer_upper[1]
        self.ws['D3'] = self.pointer_upper[2]

        self.wb.save(self.file_name)

         
    def start_webcam(self):
        if self.cameraSelect.currentText() == "0" or self.cameraSelect.currentText() == "1" or self.cameraSelect.currentText() == "2":
            self.capture = cv2.VideoCapture(int(self.cameraSelect.currentText()))
        else:
            self.capture = cv2.VideoCapture(self.cameraSelect.currentText())

        self.capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 360)
        self.capture.set(cv2.CAP_PROP_FRAME_WIDTH, 480)

        self.length = int(self.capture.get(cv2.CAP_PROP_FRAME_COUNT))
           
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame)
        self.timer.start(30)
        
    def update_frame(self):
        ret, self.image = self.capture.read()
        self.frame_counter += 1
        if self.frame_counter == self.length:
            self.frame_counter = 0
            if self.cameraSelect.currentText() == "0" or self.cameraSelect.currentText() == "1" or self.cameraSelect.currentText() == "2":
                self.capture = cv2.VideoCapture(int(self.cameraSelect.currentText()))
            else:
                self.capture = cv2.VideoCapture(self.cameraSelect.currentText())
    
        color_lower = np.array([self.h_min.value(), self.s_min.value(), self.v_min.value()], np.uint8)
        color_upper = np.array([self.h_max.value(), self.s_max.value(), self.v_max.value()], np.uint8)
        
        self.h_min_val.setText(str(self.h_min.value()))
        self.s_min_val.setText(str(self.s_min.value()))
        self.v_min_val.setText(str(self.v_min.value()))
        self.h_max_val.setText(str(self.h_max.value()))
        self.s_max_val.setText(str(self.s_max.value()))
        self.v_max_val.setText(str(self.v_max.value()))

        blur = cv2.GaussianBlur(self.image,(5,5),0)
        hsv = cv2.cvtColor(blur, cv2.COLOR_BGR2HSV)
         
        hsv_not_blur = cv2.cvtColor(self.image, cv2.COLOR_BGR2HSV)

        if self.mode == "corner":
            hsv_realtime = hsv
        else:
            hsv_realtime = hsv_not_blur

        self.color_mask = cv2.inRange(hsv_realtime, color_lower, color_upper)

        color_mask_pointer = cv2.inRange(hsv_not_blur, self.pointer_lower, self.pointer_upper)

        trackedImage = self.track_screen(self.image.copy(), color_mask_pointer, self.a1, self.a2, self.a3, self.a4)

        self.displayImage(trackedImage,2)

        self.displayImage(self.color_mask,1)
        
    def track_screen(self,img, color_mask_pointer, c1, c2, c3, c4):

        self.c1.setText(str(c1))
        self.c2.setText(str(c2))
        self.c3.setText(str(c3))
        self.c4.setText(str(c4))

        cv2.circle(img,(c1),5,(0, 255, 0),-1)
        cv2.circle(img,(c2),5,(0, 255, 0),-1)
        cv2.circle(img,(c3),5,(0, 255, 0),-1)
        cv2.circle(img,(c4),5,(0, 255, 0),-1)

        cnts = cv2.findContours(color_mask_pointer, cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)

        if len(cnts) > 0:
            c = max(cnts, key=cv2.contourArea)
            ((xp,yp), radius) = cv2.minEnclosingCircle(c)
            xp = int(xp)
            yp = int(yp)
            cv2.circle(img,(xp,yp),5,(0, 255, 255),-1)
            self.laser_pointer_val.setText(str((xp,yp))) 

            points = np.float32([[[xp, yp]]])

            if np.any(self.h != None):
                transformed = cv2.perspectiveTransform(points, self.h)
                
                xt = int(transformed[0][0][0])
                yt = int(transformed[0][0][1])

                self.count += 1
                self.last_xt = xt
                self.last_yt = yt
            
            else:
                if self.count > 1 and self.count < 50 and self.mouse_track == True and xt > 0 and xt < self.screenExtends and yt > 0 and yt < self.screenHeight:
                    pyautogui.click(x = self.last_xt, y = self.last_xt, click = 2, interval = 0.25)
                    
                xt = 0
                yt = 0
                self.count = 0

            self.mouse_pointer_val.setText(str((self.screenWidth-xt,yt)))

            if self.mouse_track == True and xt > 0 and xt < self.screenExtends and yt > 0 and yt < self.screenHeight:
                pyautogui.moveTo(self.screenWidth-xt,yt)
                #win32api.SetCursorPos((2389-xt,yt))

        return img
        
    def stop_webcam(self):
        self.capture.release()
        self.timer.stop()
    
   
    def displayImage(self,img,window=1):
        qformat = QImage.Format_Indexed8
        if len(img.shape) == 3: 
            if img.shape[2] == 4:
                qformat = QImage.Format_RGBA8888
            else:
                qformat = QImage.Format_RGB888
        
        outImage = QImage(img, img.shape[1], img.shape[0], img.strides[0],qformat)

        outImage = outImage.rgbSwapped()
        
        if window == 1:
            self.image_label1.setPixmap(QPixmap.fromImage(outImage))
            self.image_label1.setScaledContents(True)
        
        if window == 2:
            self.image_label2.setPixmap(QPixmap.fromImage(outImage))
            self.image_label2.setScaledContents(True)
        
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ColorDetector()
    window.setWindowTitle('Laser tracker')
    window.setWindowFlags(Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)
    window.show()
    sys.exit(app.exec_())